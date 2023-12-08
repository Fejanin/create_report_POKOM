import openpyxl
from my_lib.base_func import get_base_form



# Номенклатура | Ед. изм. | Начальный остаток | Приход | Расход | Конечный остаток | метка | крат | ср | заяв | разн |
# => заказ | заказ | заказ | заказ | заказ | заказ | заказ | пуд | заказ | заказ
# =>| ср | заказ | кон ост | факт | медв | тк | атпр | пудп | ср | ср | пр | коментарии | сум | вес



file = 'дв 29,11,23пок.xlsx'
res, ALL_COLUMNS = get_base_form(file)

new_file_name = 'test.xlsx'
res_wb = openpyxl.Workbook()
res_ws = res_wb.active



for r, item in enumerate(res, 1):
    if r == 3:
        for c, i in enumerate(ALL_COLUMNS, 1):
            if ALL_COLUMNS[c - 1][0] == 'Конечный остаток':
                res_ws.cell(row=r, column=c + 1).value = ALL_COLUMNS[c - 1][0]
            else:
                res_ws.cell(row = r, column = c).value = ALL_COLUMNS[c - 1][0]
    elif r > 3:
        res_ws.cell(row=r, column=1).value = item
        for c in range(1, len(ALL_COLUMNS)):
            if ALL_COLUMNS[c][0] == 'Конечный остаток':
                res_ws.cell(row=r, column=c + 2).value = res[item][ALL_COLUMNS[c][0]]
            else:
                try:
                    res_ws.cell(row=r, column=c + 1).value = round(float(res[item][ALL_COLUMNS[c][0]]))
                except:
                    res_ws.cell(row=r, column=c + 1).value = res[item][ALL_COLUMNS[c][0]]
        
    

res_wb.save(filename=new_file_name)

