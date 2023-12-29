import openpyxl
from my_lib.table import Table, Row

def read_xlsx_file(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    return wb, ws, max_col, max_row

def create_base_table_from_dv_file(file_dv):
    table = Table()
    wb, ws, max_col, max_row = read_xlsx_file(file_dv)

    all_columns = {
        'name_sku': None,  # Номенклатура
        'units_of_measurement': None,  # Ед. изм.
        'initial_balance': None,  # Начальный остаток
        'receipt_of_products': None,  # Приход
        'final_balance': None,  # Конечный остаток
    }

    flag = False
    for row in range(1, max_row + 1):
        current_name_row = None
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not all(all_columns.values()):
                if cell_value == 'Номенклатура':
                    all_columns['name_sku'] = col
                if cell_value == 'Ед. изм.':
                    all_columns['units_of_measurement'] = col
                if cell_value == 'Начальный остаток':
                    all_columns['initial_balance'] = col
                if cell_value == 'Приход':
                    all_columns['receipt_of_products'] = col
                if cell_value == 'Конечный остаток':
                    all_columns['final_balance'] = col
            if flag and col in all_columns.values():
                if col == all_columns['name_sku']:
                    current_name_row = cell_value
                    table.add_row(Row(current_name_row))
                elif current_name_row and col == all_columns['units_of_measurement']:
                    table.rows[table.rows.index(current_name_row)].units_of_measurement = cell_value
                elif current_name_row and col == all_columns['initial_balance']:
                    table.rows[table.rows.index(current_name_row)].initial_balance = cell_value
                elif current_name_row and col == all_columns['receipt_of_products']:
                    table.rows[table.rows.index(current_name_row)].receipt_of_products = cell_value
                elif current_name_row and col == all_columns['final_balance']:
                    table.rows[table.rows.index(current_name_row)].final_balance = cell_value
        if ws.cell(row=row, column=1).value == 'Основной склад ПОКОМ':
            flag = True
        if ws.cell(row=row + 1, column=1).value == 'Склад корректировок':
            flag = False
    return table


def add_data_from_old_file(file_dv_old, table):
    wb, ws, max_col, max_row = read_xlsx_file(file_dv_old)

    all_columns = {
        'name_sku': None,  # Номенклатура
        'units_of_measurement': None,  # Ед. изм.,
        'mark': None,  # метка
        'multiplicity': None,  # крат
        'expiration_dates': None,  # сроки
        'orders_is_on_the_way': {},  # заказ в пути
        'medvedev_sales': None,  # медв
        'tk_sales': None,  # тк
        'atamanov_sales': None,  # атпр
        'pud_sales': None,  # пудп
        'average_values': {},  # ср: {'10,11,': 28, ...}
        'comments': None,  # комментарии
    }

    columns_val = []
    columns_dict = []

    for row in range(1, max_row + 1):
        current_sku = None
        current_obj = None
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not all(all_columns.values()):
                if cell_value == 'Номенклатура':
                    all_columns['name_sku'] = col
                    columns_val.append(col)
                if cell_value == 'Ед. изм.':
                    all_columns['units_of_measurement'] = col
                    columns_val.append(col)
                if cell_value == 'метка':
                    all_columns['mark'] = col
                    columns_val.append(col)
                if cell_value == 'крат':
                    all_columns['multiplicity'] = col
                    columns_val.append(col)
                if cell_value == 'сроки':
                    all_columns['expiration_dates'] = col
                    columns_val.append(col)
                if cell_value == 'заказ в пути':
                    if ws.cell(row=row + 1, column=col).value:
                        all_columns['orders_is_on_the_way'].update({ws.cell(row=row + 1, column=col).value: col})
                        columns_dict.append(col)
                if cell_value == 'медв':
                    all_columns['medvedev_sales'] = col
                    columns_val.append(col)
                if cell_value == 'тк':
                    all_columns['tk_sales'] = col
                    columns_val.append(col)
                if cell_value == 'атпр':
                    all_columns['atamanov_sales'] = col
                    columns_val.append(col)
                if cell_value == 'пудп':
                    all_columns['pud_sales'] = col
                    columns_val.append(col)
                if cell_value == 'ср':
                    if ws.cell(row=row + 1, column=col).value:
                        all_columns['average_values'].update({ws.cell(row=row + 1, column=col).value: col})
                        columns_dict.append(col)
                if cell_value == 'комментарии':
                    all_columns['comments'] = col
                    columns_val.append(col)
            if all(all_columns.values()) and col == all_columns['name_sku'] and cell_value:
                current_sku = cell_value
                if not current_sku in table.rows:  # Добавить СКЮ, которого нет в новом движении
                    table.add_row(Row(current_sku))
                current_obj = table.rows[table.rows.index(current_sku)]
            if all(all_columns.values()) and current_sku and col in columns_val:
                name_col = list(filter(lambda x: x[1] == col, all_columns.items()))[0][0]
                if name_col == 'name_sku':
                    continue
                current_obj.__dict__[name_col] = cell_value
            elif all(all_columns.values()) and current_sku and col in columns_dict:
                if col in all_columns['orders_is_on_the_way'].values():
                    value = cell_value
                    if value == '#REF!':
                        value = 0
                    key_or_in_way = list(filter(lambda x: x[1] == col, all_columns['orders_is_on_the_way'].items()))[0][0]
                    current_obj.__dict__['orders_is_on_the_way'][key_or_in_way] = value
                    table.COLUMNS['orders_is_on_the_way'][1].update({key_or_in_way: None})
                elif col in all_columns['average_values'].values():
                    value = ws.cell(row=row, column=col).value
                    key_av_values = list(filter(lambda x: x[1] == col, all_columns['average_values'].items()))[0][0]
                    current_obj.__dict__['average_values'][key_av_values] = value
                    table.COLUMNS['average_values'][1].update({key_av_values: None})
                else:
                    raise Exception('NOT FOUND!!!')


def add_data_from_sales_file(file, table, names):
    columns = {'Номенклатура': None}
    columns[names[0]] = None

    wb, ws, max_col, max_row = read_xlsx_file(file)
    for row in range(1, max_row + 1):
        current_obj = None
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value in columns:  # находим колонки: Номенклатура и количество
                if not columns[cell_value]:
                    columns[cell_value] = col  # связываем название колонки с ее номером
            if col in columns.values():
                if cell_value in table.rows:
                    current_obj = table.rows[table.rows.index(cell_value)]
                elif current_obj and cell_value:
                    current_obj.__dict__[names[1]] = cell_value
